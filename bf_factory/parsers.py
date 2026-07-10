"""常用BF文件的基础解析和来源定位。"""

import csv
import io
import json
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from .storage import sanitize_filename


MAX_UPLOAD_BYTES = 30 * 1024 * 1024
SUPPORTED_EXTENSIONS = {
    ".doc", ".docx", ".ppt", ".pptx", ".pdf",
    ".jpg", ".jpeg", ".png", ".webp",
    ".xls", ".xlsx", ".csv", ".txt", ".md",
}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
OOXML_ROOTS = {".docx": "word/document.xml", ".pptx": "ppt/presentation.xml", ".xlsx": "xl/workbook.xml"}
OLE_MAGIC = bytes.fromhex("D0CF11E0A1B11AE1")


class BFParseError(ValueError):
    pass


def validate_upload(filename, data, max_bytes=MAX_UPLOAD_BYTES):
    name = sanitize_filename(filename)
    extension = Path(name).suffix.lower()
    payload = bytes(data or b"")
    if extension not in SUPPORTED_EXTENSIONS:
        raise BFParseError(f"不支持的BF文件格式: {extension or '无扩展名'}")
    if not payload:
        raise BFParseError("BF文件不能为空")
    if len(payload) > max_bytes:
        raise BFParseError(f"BF文件超过{max_bytes // 1024 // 1024}MB限制")
    if payload.startswith(b"MZ"):
        raise BFParseError("文件头与声明格式不匹配")
    if extension in OOXML_ROOTS:
        try:
            with zipfile.ZipFile(io.BytesIO(payload)) as archive:
                if OOXML_ROOTS[extension] not in archive.namelist() and extension != ".pptx":
                    raise BFParseError("Office文件结构无效")
                if extension == ".pptx" and not any(name.startswith("ppt/slides/slide") for name in archive.namelist()):
                    raise BFParseError("PPT文件缺少页面")
        except zipfile.BadZipFile as exc:
            raise BFParseError("Office文件结构无效") from exc
    elif extension in {".doc", ".ppt", ".xls"} and not payload.startswith(OLE_MAGIC):
        raise BFParseError("旧版Office文件头无效")
    elif extension == ".pdf" and not payload.startswith(b"%PDF"):
        raise BFParseError("PDF文件头无效")
    elif extension == ".png" and not payload.startswith(b"\x89PNG\r\n\x1a\n"):
        raise BFParseError("PNG文件头无效")
    elif extension in {".jpg", ".jpeg"} and not payload.startswith(b"\xff\xd8\xff"):
        raise BFParseError("JPEG文件头无效")
    elif extension == ".webp" and not (payload.startswith(b"RIFF") and payload[8:12] == b"WEBP"):
        raise BFParseError("WebP文件头无效")
    elif extension in {".txt", ".md", ".csv"}:
        if b"\x00" in payload[:4096]:
            raise BFParseError("文本文件包含二进制内容")
        _decode_text(payload)
    return {"filename": name, "extension": extension, "sizeBytes": len(payload)}


def parse_document(filename, data):
    info = validate_upload(filename, data)
    extension = info["extension"]
    if extension in {".txt", ".md"}:
        return _parse_text(info["filename"], data, extension)
    if extension == ".csv":
        return _parse_csv(info["filename"], data)
    if extension == ".docx":
        return _parse_docx(info["filename"], data)
    if extension == ".pptx":
        return _parse_pptx(info["filename"], data)
    if extension == ".pdf":
        return _parse_pdf(info["filename"], data)
    if extension in IMAGE_EXTENSIONS:
        return _parse_image(info["filename"], data, extension)
    if extension == ".xlsx":
        return _parse_xlsx(info["filename"], data)
    if extension in {".doc", ".ppt", ".xls"}:
        return _parse_legacy_office(info["filename"], data, extension)
    raise BFParseError("暂不支持该BF文件")


def _segment(*, text="", block_type="TEXT", page_no=None, slide_no=None, paragraph_no=None, sheet_name=None, cell_range=None, table=None, locator=None, confidence=None):
    return {
        "text": str(text or "").strip(),
        "blockType": block_type,
        "pageNo": page_no,
        "slideNo": slide_no,
        "paragraphNo": paragraph_no,
        "sheetName": sheet_name,
        "cellRange": cell_range,
        "table": table or [],
        "locator": locator or {},
        "confidence": confidence,
    }


def _result(file_format, filename, segments, warnings=None):
    return {
        "format": file_format,
        "filename": filename,
        "segments": [item for item in segments if item["text"] or item["table"] or item["blockType"] == "IMAGE"],
        "warnings": list(warnings or []),
    }


def _decode_text(data):
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return bytes(data).decode(encoding)
        except UnicodeDecodeError:
            continue
    raise BFParseError("无法识别文本编码")


def _parse_text(filename, data, extension):
    lines = _decode_text(data).splitlines()
    segments = [
        _segment(text=line, paragraph_no=index, locator={"paragraphNo": index})
        for index, line in enumerate(lines, 1)
        if line.strip()
    ]
    return _result("MARKDOWN" if extension == ".md" else "TEXT", filename, segments)


def _parse_csv(filename, data):
    rows = list(csv.reader(io.StringIO(_decode_text(data))))
    end_column = _column_name(max((len(row) for row in rows), default=1))
    cell_range = f"A1:{end_column}{max(len(rows), 1)}"
    return _result(
        "CSV",
        filename,
        [_segment(block_type="TABLE", sheet_name="CSV", cell_range=cell_range, table=rows, locator={"sheetName": "CSV", "cellRange": cell_range})],
    )


def _parse_docx(filename, data):
    segments = []
    warnings = []
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        root = ET.fromstring(archive.read("word/document.xml"))
        body = root.find("w:body", ns)
        paragraph_no = 0
        if body is not None:
            for child in list(body):
                local = child.tag.rsplit("}", 1)[-1]
                if local == "p":
                    paragraph_no += 1
                    text = "".join(node.text or "" for node in child.findall(".//w:t", ns)).strip()
                    if text:
                        segments.append(_segment(text=text, paragraph_no=paragraph_no, locator={"paragraphNo": paragraph_no}))
                elif local == "tbl":
                    rows = []
                    for row in child.findall(".//w:tr", ns):
                        rows.append(["".join(node.text or "" for node in cell.findall(".//w:t", ns)).strip() for cell in row.findall("w:tc", ns)])
                    segments.append(_segment(block_type="TABLE", paragraph_no=paragraph_no + 1, table=rows, locator={"afterParagraph": paragraph_no}))
        for media_name in sorted(name for name in archive.namelist() if name.startswith("word/media/") and not name.endswith("/")):
            ocr_text, ocr_warning = _ocr_image_bytes(archive.read(media_name))
            segments.append(_segment(text=ocr_text, block_type="IMAGE", locator={"embeddedPath": media_name, "ocrRequired": not bool(ocr_text)}))
            if ocr_warning:
                warnings.append(f"{media_name}: {ocr_warning}")
    return _result("DOCX", filename, segments, warnings)


def _parse_pptx(filename, data):
    segments = []
    warnings = []
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        slide_names = sorted(
            (name for name in archive.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)),
            key=_numeric_suffix,
        )
        for slide_no, slide_name in enumerate(slide_names, 1):
            root = ET.fromstring(archive.read(slide_name))
            texts = [node.text.strip() for node in root.findall(".//a:t", ns) if node.text and node.text.strip()]
            if texts:
                segments.append(_segment(text="\n".join(texts), block_type="SLIDE", page_no=slide_no, slide_no=slide_no, locator={"slideNo": slide_no}))
            notes_name = f"ppt/notesSlides/notesSlide{slide_no}.xml"
            if notes_name in archive.namelist():
                notes_root = ET.fromstring(archive.read(notes_name))
                notes = [node.text.strip() for node in notes_root.findall(".//a:t", ns) if node.text and node.text.strip()]
                if notes:
                    segments.append(_segment(text="\n".join(notes), block_type="NOTES", page_no=slide_no, slide_no=slide_no, locator={"slideNo": slide_no, "kind": "notes"}))
        for media_name in sorted(name for name in archive.namelist() if name.startswith("ppt/media/") and not name.endswith("/")):
            ocr_text, ocr_warning = _ocr_image_bytes(archive.read(media_name))
            segments.append(_segment(text=ocr_text, block_type="IMAGE", locator={"embeddedPath": media_name, "ocrRequired": not bool(ocr_text)}))
            if ocr_warning:
                warnings.append(f"{media_name}: {ocr_warning}")
    return _result("PPTX", filename, segments, warnings)


def _parse_pdf(filename, data):
    segments = []
    warnings = []
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data), strict=False)
        for page_no, page in enumerate(reader.pages, 1):
            text = str(page.extract_text() or "").strip()
            if text:
                segments.append(_segment(text=text, block_type="PAGE", page_no=page_no, locator={"pageNo": page_no}))
    except Exception as exc:
        warnings.append(f"PDF标准解析降级: {type(exc).__name__}")
    if not segments:
        text_chunks = []
        for raw in re.findall(rb"\(([^()]*)\)\s*Tj", bytes(data)):
            text_chunks.append(_decode_pdf_literal(raw))
        text = "\n".join(item for item in text_chunks if item.strip()).strip()
        if text:
            segments.append(_segment(text=text, block_type="PAGE", page_no=1, locator={"pageNo": 1, "parseMode": "literal-fallback"}))
        else:
            segments.append(_segment(block_type="IMAGE", page_no=1, locator={"pageNo": 1, "ocrRequired": True, "parseMode": "scanned-pdf"}))
            warnings.append("扫描型PDF需要OCR或MMN多模态识别")
    return _result("PDF", filename, segments, warnings)


def _parse_image(filename, data, extension):
    text, warning = _ocr_image_bytes(data)
    return _result(
        extension[1:].upper(),
        filename,
        [_segment(text=text, block_type="IMAGE", page_no=1, locator={"pageNo": 1, "ocrRequired": not bool(text)})],
        [warning] if warning else [],
    )


def _parse_xlsx(filename, data):
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        segments = []
        for sheet in workbook.worksheets:
            rows = [["" if value is None else str(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            while rows and not any(cell for cell in rows[-1]):
                rows.pop()
            end_column = _column_name(max((len(row) for row in rows), default=1))
            cell_range = f"A1:{end_column}{max(len(rows), 1)}"
            segments.append(_segment(block_type="TABLE", sheet_name=sheet.title, cell_range=cell_range, table=rows, locator={"sheetName": sheet.title, "cellRange": cell_range}))
        return _result("XLSX", filename, segments)
    except Exception as exc:
        raise BFParseError(f"Excel解析失败: {type(exc).__name__}") from exc


def _parse_legacy_office(filename, data, extension):
    soffice = shutil.which("soffice")
    if not soffice:
        raise BFParseError("旧版Office解析需要LibreOffice")
    target_extension = {".doc": ".docx", ".ppt": ".pptx", ".xls": ".xlsx"}[extension]
    with tempfile.TemporaryDirectory(prefix="mmn-bf-office-") as tmp:
        source = Path(tmp) / sanitize_filename(filename)
        source.write_bytes(data)
        process = subprocess.run(
            [soffice, "--headless", "--convert-to", target_extension[1:], "--outdir", tmp, str(source)],
            capture_output=True,
            timeout=45,
            check=False,
        )
        converted = source.with_suffix(target_extension)
        if process.returncode != 0 or not converted.exists():
            raise BFParseError("旧版Office转换失败")
        result = parse_document(converted.name, converted.read_bytes())
        result["warnings"].append(f"原始{extension}已通过LibreOffice转换为{target_extension}")
        result["originalFormat"] = extension[1:].upper()
        return result


def _ocr_image_bytes(data):
    try:
        from PIL import Image
        import pytesseract

        text = pytesseract.image_to_string(Image.open(io.BytesIO(data)), lang="chi_sim+eng").strip()
        return text, "" if text else "OCR未识别到文字"
    except ImportError:
        tesseract = shutil.which("tesseract")
        if not tesseract:
            return "", "本地OCR组件未安装，等待MMN多模态识别"
        try:
            languages = subprocess.run([tesseract, "--list-langs"], capture_output=True, text=True, timeout=10, check=False).stdout
            language = "chi_sim+eng" if "chi_sim" in languages else "eng"
            result = subprocess.run(
                [tesseract, "stdin", "stdout", "-l", language],
                input=data,
                capture_output=True,
                timeout=30,
                check=False,
            )
            text = result.stdout.decode("utf-8", errors="replace").strip()
            return text, "" if text else "OCR未识别到文字"
        except Exception as exc:
            return "", f"本地OCR降级: {type(exc).__name__}"
    except Exception as exc:
        return "", f"本地OCR降级: {type(exc).__name__}"


def _numeric_suffix(path):
    match = re.search(r"(\d+)\.xml$", path)
    return int(match.group(1)) if match else 0


def _decode_pdf_literal(raw):
    text = raw.replace(rb"\(", b"(").replace(rb"\)", b")").replace(rb"\\", b"\\")
    for encoding in ("utf-8", "latin-1"):
        try:
            return text.decode(encoding)
        except UnicodeDecodeError:
            continue
    return ""


def _column_name(number):
    number = max(int(number), 1)
    result = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        result = chr(65 + remainder) + result
    return result
